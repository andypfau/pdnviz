from .power_base import PowerConfig, PowerBaseElement
from .power_component import PowerComponent, PowerHierarchy, PowerHierarchyElement, GroupSource, PowerInput, PowerOutput
from .power_sources import Supply, MultiSupply, DualSupply
from .power_sinks import Load, MultiLoad, DualLoad
from .power_converters import LDO, DcDc
from .systools import get_tempfile_path, open_file, ensure_directories
import openpyxl, datetime, os, subprocess, tempfile


class PowerSpreadsheet:


    def __init__(self, pdn: "PowerComponent", title: str = None, grouped: bool = False):

        if title is None: title = 'PDN'
        
        self.title, self.grouped = title, grouped

        self._wb = openpyxl.Workbook()
        self.initial_sheets = self._wb.sheetnames
        self._define_formats()
        self._problems = []
        self._sheets = {}
        self._title_sheet = None
        
        hierarchies = pdn.get_hierarchy(grouped=self.grouped)
        self.all_groups = [h.group for h in hierarchies]
        for hierarchy in hierarchies:
            self._add_sheet(hierarchy)
        
        if len(self._sheets) < 1:
            raise RuntimeError('No components were found to put into the spreadsheet')
        
        self._finalize()
    

    def save(self, path: str, view: bool = False, makedirs: bool = True):
        if makedirs:
            ensure_directories(path, is_filename=True)
        self._wb.save(path)
        if view:
            open_file(path)
    

    def view(self) -> str:
        path = get_tempfile_path()
        self._wb.save(path)
        open_file(path)
        return path
    

    def _add_sheet(self, hierarchy: PowerHierarchy):
        if len(hierarchy.all_dissipating_components) < 1:
            return # probably an empty group?
        title = self._group_name(hierarchy.group) if self.grouped else self.title
        sheet = self._wb.create_sheet(title=self._human_name(title))
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 11
        sheet.column_dimensions['D'].width = 11
        sheet.column_dimensions['E'].width = 11
        sheet.column_dimensions['F'].width = 11
        sheet.column_dimensions['G'].width = 11
        sheet.column_dimensions['H'].width = 11
        sheet.column_dimensions['I'].width = 11
        sheet.column_dimensions['J'].width = 11
        sheet.column_dimensions['K'].width = 30
        sheet.column_dimensions['L'].width = 30
        row = 1
        sheet.cell(row,1).value = title
        sheet.cell(row,1).font = openpyxl.styles.Font(bold=True, underline='single')
        row += 2
        summary = []

        args = (hierarchy, sheet)
        state = (row, summary)
        state = self._add_drawn_power_table(*args, *state)
        #state = self._add_converter_table(*args, *state)
        state = self._add_hierarchy_table(*args, *state)
        state = self._add_dissipation_table(*args, *state)
        state = self._add_sourced_power_table(*args, *state)
        (row, summary) = state
        
        if len(self.all_groups)<=1: # if there are multiple groups, we place the timestamp in the summary
            row += 1
            self._add_timestamp(sheet, row)

        self._sheets[sheet.title] = ', '.join(summary)


    def _add_drawn_power_table(self, hierarchy: PowerHierarchy, sheet: "openpyxl.worksheet.worksheet.Worksheet", row: int, summary: "list[str]"):
        if len(hierarchy.sources) <= 0:
            return (row, summary)
        sheet.cell(row,1).value = 'Power Drawn From Sources'
        sheet.cell(row,1).font = openpyxl.styles.Font(underline='single')
        row += 2
        p_tot = 0
        sheet.cell(row,1).value = 'Source'
        sheet.cell(row,2).value = 'Voltage'
        sheet.cell(row,3).value = 'I Drawn'
        sheet.cell(row,4).value = 'P Drawn'
        sheet.cell(row,5).value = 'Warnings'
        row += 1
        row1 = row
        for source in hierarchy.sources:
            name = source.output.full_name(include_group=True)
            if source.output.parent.group != hierarchy.group:
                name += f' ({source.output.parent.group})'
            sheet.cell(row,1).value = name
            sheet.cell(row,2).value = source.output.v_out
            sheet.cell(row,3).value = source.i_drawn
            sheet.cell(row,4).value = source.p_drawn
            
            if source.output.i_out_max is not None:
                sheet.cell(row,3).comment = openpyxl.comments.Comment(f'max. I_out: {source.output.i_out_max:.3g} A', 'PDN Viz')
            if source.output.p_out_max is not None:
                sheet.cell(row,4).comment = openpyxl.comments.Comment(f'max. P_out: {source.output.p_out_max:.3g} W', 'PDN Viz')
                
            if not source.output.ok():
                sheet.cell(row, 1).font = openpyxl.styles.Font(color='FF0000') # also make the 1st col red (easier visible)
                sheet.cell(row, 5).value = '; '.join(source.output.get_warnings())
                sheet.cell(row, 5).font = openpyxl.styles.Font(color='FF0000')
                for p in source.output.get_warnings():
                    self._problems.append((p,sheet.title,row))
            
            p_tot += source.p_drawn
            sheet.cell(row,2).number_format = '0.###" V"'
            sheet.cell(row,3).number_format = '0.###" A"'
            sheet.cell(row,4).number_format = '0.###" W"'
            row2 = row
            row += 1
        row += 1 # add emtpy row, otherwise the sum will be part of the table, which messes up sorting
        sheet.cell(row,4).value = f'=SUM({self._range(row1, row2, 4, 4)})'
        sheet.cell(row,4).number_format = '0.###" W"'
        sheet.cell(row,4).font = openpyxl.styles.Font(underline='double')
        self._newtable(sheet, self._range(row1-1, row2, 1, 5), self._code_name(self._group_name(hierarchy.group))+'_Drawn')
        sheet.conditional_formatting.add(self._range(row1, row2, 2, 2), self._data_bar_v)
        sheet.conditional_formatting.add(self._range(row1, row2, 3, 3), self._data_bar_i)
        sheet.conditional_formatting.add(self._range(row1, row2, 4, 4), self._data_bar_p)
        row += 2
        summary.append(f'draws {p_tot:.3g} W')
        return (row, summary)


    def _add_hierarchy_table(self, hierarchy: PowerHierarchy, sheet: "openpyxl.worksheet.worksheet.Worksheet", row: int, summary: "list[str]"):
        
        if len(hierarchy.sink_hierarchy) <= 0:
            return (row, summary)
        
        sheet.cell(row,1).value = 'Component Supply Hierarchy'
        sheet.cell(row,1).font = openpyxl.styles.Font(underline='single')
        row += 2
        sheet.cell(row,1).value = 'Component'
        sheet.cell(row,2).value = 'Source'
        sheet.cell(row,3).value = 'V In'
        sheet.cell(row,4).value = 'I Drawn'
        sheet.cell(row,5).value = 'P Drawn'
        sheet.cell(row,6).value = 'Notes'
        sheet.cell(row,7).value = 'Warnings'
        row += 1
        
        row1 = row
        row2 = row

        def print_level(hierarchy_level: "list[PowerHierarchyElement]", level_index: int):
            nonlocal row, hierarchy, row2
            group_row1 = row
            group_row2 = None

            if level_index==0:
                indent = ''
            elif level_index==1:
                indent = '   ⮩ '
            else:
                indent = '   '*level_index + '   ⮩ '

            for level in hierarchy_level:
                power_input = level.input
                sheet.cell(row,1).value = indent + power_input.full_name()
                sheet.cell(row,2).value = power_input.source.full_name()
                sheet.cell(row,3).value = power_input.v_in_actual
                sheet.cell(row,3).number_format = '0.###" V"'
                sheet.cell(row,4).value = power_input.i_in*1e3
                sheet.cell(row,4).number_format = '0" mA"'
                comment = f'total I_out of source: {power_input.source.i_out_calc:.3g} A'
                if power_input.source.i_out_max is not None:
                    comment += f'\nmax. I_out of source: {power_input.source.i_out_max:.3g} A'
                sheet.cell(row,4).comment = openpyxl.comments.Comment(comment, 'PDN Viz')
                sheet.cell(row,5).value = power_input.p_in_calc*1e3
                sheet.cell(row,5).number_format = '0" mW"'
                comment = f'total P_out of source: {power_input.source.p_out_calc:.3g} W'
                if power_input.source.p_out_max is not None:
                    comment += f'\nmax. P_out of source: {power_input.source.p_out_max:.3g} W'
                sheet.cell(row,5).comment = openpyxl.comments.Comment(comment, 'PDN Viz')
                
                if isinstance(power_input.parent, LDO):
                    sheet.cell(row,6).value = f'{power_input.parent.v_drop_calc:.3g} V drop'
                    if power_input.parent.v_drop_min is not None:
                        sheet.cell(row,6).comment = openpyxl.comments.Comment(f'min. req. drop: {power_input.parent.v_drop_min:.3g} V', 'PDN Viz')
                if isinstance(power_input.parent, DcDc):
                    sheet.cell(row,6).value = f'{power_input.parent.eff_pct_calc:.0f}% eff.'
                
                if not power_input.parent.ok():
                    sheet.cell(row, 1).font = openpyxl.styles.Font(color='FF0000') # also make the 1st col red (easier visible)
                    sheet.cell(row, 7).value = '; '.join(power_input.parent.get_warnings())
                    sheet.cell(row, 7).font = openpyxl.styles.Font(color='FF0000')
                    for p in power_input.parent.get_warnings():
                        self._problems.append((p,sheet.title,row))

                row2 = row
                group_row2 = row
                row += 1

                if len(level.connected_sinks)>0:
                    print_level(level.connected_sinks, level_index+1)
                
            if (level_index>0) and (group_row2 is not None):
                sheet.row_dimensions.group(group_row1, group_row2, outline_level=level_index)
        
        print_level(hierarchy.sink_hierarchy, 0)
        row += 1 # add emtpy row, otherwise the sum will be part of the table, which messes up sorting
        self._newtable(sheet, self._range(row1-1, row2, 1, 7), self._code_name(self._group_name(hierarchy.group))+'_Hier')
        sheet.conditional_formatting.add(self._range(row1, row2, 3, 3), self._data_bar_v)
        sheet.conditional_formatting.add(self._range(row1, row2, 4, 4), self._data_bar_i)

        row += 2
        return (row, summary)


    def _add_dissipation_table(self, hierarchy: PowerHierarchy, sheet: "openpyxl.worksheet.worksheet.Worksheet", row: int, summary: "list[str]"):
        if len(hierarchy.all_dissipating_components) <= 0:
            return (row, summary)
        
        sheet.cell(row,1).value = 'Power-Dissipating Components'
        sheet.cell(row,1).font = openpyxl.styles.Font(underline='single')
        row += 2
        sheet.cell(row,1).value = 'Component'
        sheet.cell(row,2).value = 'Sources'
        sheet.cell(row,3).value = 'P Dissipated'
        sheet.cell(row,4).value = 'T J'
        sheet.cell(row,5).value = 'Warnings'
        row += 1
        
        p_diss_tot = 0
        
        row1 = row
        row2 = row
        
        for component in hierarchy.all_dissipating_components:
            
            sheet.cell(row,1).value = component.name
            sheet.cell(row,2).value = ', '.join([s.full_name() for s in component._inputs])
            
            sheet.cell(row,3).value = component.p_diss_calc*1e3
            sheet.cell(row,3).number_format = '0" mW"'
            p_diss_tot += component.p_diss_calc
            if component.p_diss_max is not None:
                sheet.cell(row,3).comment = openpyxl.comments.Comment(f'max. P_diss: {component.p_diss_max:.3g} W', 'PDN Viz')
            
            sheet.cell(row,4).value = component.t_j_calc
            sheet.cell(row,4).number_format = '0.#" °C"'
            if component.t_j_max is not None and component.r_th_ja is not None:
                sheet.cell(row,4).comment = openpyxl.comments.Comment(f'R_Th_JA: {component.r_th_ja:.1f} K/W\nmax. T_J: {component.t_j_max:.1f} °C', 'PDN Viz')
                
            if not component.ok():
                sheet.cell(row, 1).font = openpyxl.styles.Font(color='FF0000') # also make the 1st col red (easier visible)
                sheet.cell(row, 5).value = '; '.join(component.get_warnings())
                sheet.cell(row, 5).font = openpyxl.styles.Font(color='FF0000')
                for p in component.get_warnings():
                    self._problems.append((p,sheet.title,row))

            row2 = row
            row += 1

        row += 1 # add emtpy row, otherwise the sum will be part of the table, which messes up sorting
        sheet.cell(row,3).value = f'=SUM({self._range(row1, row2, 3, 3)})/1e3'
        sheet.cell(row,3).number_format = '0.###" W"'
        sheet.cell(row,3).font = openpyxl.styles.Font(underline='double')
        self._newtable(sheet, self._range(row1-1, row2, 1, 5), self._code_name(self._group_name(hierarchy.group))+'_Comps')
        sheet.conditional_formatting.add(self._range(row1, row2, 3, 3), self._data_bar_p)
        sheet.conditional_formatting.add(self._range(row1, row2, 4, 4), self._data_bar_t)
        row += 1

        row += 1
        summary.append(f'dissipates {p_diss_tot:.3g} W')
        return (row, summary)



    def _add_sourced_power_table(self, hierarchy: PowerHierarchy, sheet: "openpyxl.worksheet.worksheet.Worksheet", row: int, summary: "list[str]"):
        if len(hierarchy.to_external) <= 0:
            return (row, summary)
    
        p_prov_tot = 0
        sheet.cell(row,1).value = 'Power Provided to External'
        sheet.cell(row,1).font = openpyxl.styles.Font(underline='single')
        row += 2
        sheet.cell(row,1).value = 'Source'
        sheet.cell(row,2).value = 'Sinks'
        sheet.cell(row,3).value = 'V Provided'
        sheet.cell(row,4).value = 'I Provided'
        sheet.cell(row,5).value = 'P Provided'
        row += 1
        row1 = row
        for source_to_ext in hierarchy.to_external:
            p_prov_tot += source_to_ext.p_provided
            sheet.cell(row,1).value = source_to_ext.output.full_name()
            sheet.cell(row,2).value = ', '.join(source_to_ext.receiving_groups)
            sheet.cell(row,3).value = source_to_ext.output.v_out
            sheet.cell(row,3).number_format = '0.###" V"'
            sheet.cell(row,4).value = source_to_ext.i_provided
            sheet.cell(row,4).number_format = '0.###" A"'
            if source_to_ext.output.i_out_max is not None:
                sheet.cell(row,4).comment = openpyxl.comments.Comment(f'max. I_out: {source_to_ext.output.i_out_max:.3g} A', 'PDN Viz')
            sheet.cell(row,5).value = source_to_ext.p_provided
            sheet.cell(row,5).number_format = '0.###" W"'
            if source_to_ext.output.p_out_max is not None:
                sheet.cell(row,5).comment = openpyxl.comments.Comment(f'max. P_out: {source_to_ext.output.p_out_max:.3g} W', 'PDN Viz')
            row2 = row
            row += 1
        row += 1 # add emtpy row, otherwise the sum will be part of the table, which messes up sorting
        sheet.cell(row,5).value = f'=SUM({self._range(row1, row2, 5, 5)})'
        sheet.cell(row,5).number_format = '0.###" W"'
        sheet.cell(row,5).font = openpyxl.styles.Font(underline='double')
        self._newtable(sheet, self._range(row1-1, row2, 1, 5), self._code_name(self._group_name(hierarchy.group))+'_Provided')
        sheet.conditional_formatting.add(self._range(row1, row2, 3, 3), self._data_bar_v)
        sheet.conditional_formatting.add(self._range(row1, row2, 4, 4), self._data_bar_i)
        sheet.conditional_formatting.add(self._range(row1, row2, 5, 5), self._data_bar_p)
        row += 2
        summary.append(f'proivdes {p_prov_tot:.3g} W')
        return (row, summary)


    def _add_timestamp(self, sheet, row):
        sheet.cell(row,1).value = 'Timestamp'
        sheet.cell(row,2).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')


    def _finalize(self):

        # remove any dummy sheets
        for sheet in self.initial_sheets:
            self._wb.remove(self._wb[sheet])
        self.initial_sheets = []
        self._wb.active = 0

        if self.grouped:
            # create summary
            sheet = self._wb.create_sheet(title='Summary')
            self._title_sheet = sheet
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 50
            row = 1
            sheet.cell(row,1).value = 'Summary'
            sheet.cell(row,1).font = openpyxl.styles.Font(bold=True, underline='single')
            row += 2
            for sheetname,summary in self._sheets.items():
                sheet.cell(row,1).value = sheetname
                sheet.cell(row,2).value = summary
                sheet.cell(row,2).hyperlink = f'#{sheetname}'
                row += 1
            if len(self._problems)>0:
                row += 1
                for (text,sheetname,rownum) in self._problems:
                    sheet.cell(row,1).value = 'Problem'
                    sheet.cell(row,1).font = openpyxl.styles.Font(color='FF0000')
                    sheet.cell(row,2).value = text
                    sheet.cell(row,2).hyperlink = f'#{sheetname}!A{rownum}'
                    row += 1
            row += 1
            self._add_timestamp(sheet, row)
        
        sheet = self._wb._sheets.pop(len(self._wb._sheets)-1)
        self._wb._sheets.insert(0, sheet)

    
    def _define_formats(self):
        self._table_style = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
        self._data_bar_v = openpyxl.formatting.rule.Rule(type='dataBar', dataBar=openpyxl.formatting.rule.DataBar( \
            cfvo=[openpyxl.formatting.rule.FormatObject(type='min'), openpyxl.formatting.rule.FormatObject(type='max')], \
            color="CCCCFF", minLength=None, maxLength=None))
        self._data_bar_i = openpyxl.formatting.rule.Rule(type='dataBar', dataBar=openpyxl.formatting.rule.DataBar( \
            cfvo=[openpyxl.formatting.rule.FormatObject(type='min'), openpyxl.formatting.rule.FormatObject(type='max')], \
            color="CCCCCC", minLength=None, maxLength=None))
        self._data_bar_p = openpyxl.formatting.rule.Rule(type='dataBar', dataBar=openpyxl.formatting.rule.DataBar( \
            cfvo=[openpyxl.formatting.rule.FormatObject(type='min'), openpyxl.formatting.rule.FormatObject(type='max')], \
            color="FF0000", minLength=None, maxLength=None))
        self._data_bar_t = openpyxl.formatting.rule.Rule(type='dataBar', dataBar=openpyxl.formatting.rule.DataBar( \
            cfvo=[openpyxl.formatting.rule.FormatObject(type='min'), openpyxl.formatting.rule.FormatObject(type='max')], \
            color="660066", minLength=None, maxLength=None))
    

    def _group_name(self, name: "str|Ellipsis") -> str:
        if name is None or name is ...:
            return 'Ungrouped'
        return name


    def _human_name(self, name: str):
        result = ''
        for c in name:
            if c.lower() in ' abcdefghijklmnopqrstuvwxyz0123456789+-_,;.!\'':
                result += c
            else:
                if c==':' or c=='/' or c=='\\' or c=='~':
                    result += '-'
                elif c=='"' or c=='`' or c=='´':
                    result += '\''
                else:
                    result += '_'
        return result


    def _code_name(self, name: str):
        result = ''
        for c in name:
            if c.lower() in '_abcdefghijklmnopqrstuvwxyz0123456789':
                result += c
            else:
                result += '_'
        if result[0] in '0123456789':
            result = '_' + result
        return result


    def _cell(self, row, col):
        return f'{openpyxl.utils.get_column_letter(col)}{row}'


    def _range(self, row1, row2, col1, col2):
        return f'{openpyxl.utils.get_column_letter(col1)}{row1}:{openpyxl.utils.get_column_letter(col2)}{row2}'


    def _newtable(self, sheet, ref, name):
        table = openpyxl.worksheet.table.Table(ref=ref, displayName=name, tableStyleInfo=self._table_style)
        sheet.add_table(table)
